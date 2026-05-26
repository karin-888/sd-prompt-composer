#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""Import the user's personal prompt-dictionary .xlsx files (Google Drive) into YAML.

The workbooks use a mix of formats; this script handles all of them:

1. ``        en_tag: 日本語訳`` lines stored as a single-cell row (the most common
   format used by the numbered files 1.〜21.xlsx).
2. Multi-column table rows where one cell holds an english tag and another
   holds a Japanese label (e.g. ``[区分, 項目, 表現内容, プロンプト]`` in
   ``pronpt.xlsx``'s "エロ" sheet, or 3-column ``[blank, english, japanese]``
   in ``髪型.xlsx``).
3. Inline cells like ``en, ja`` separated by a full-width space (``　``).

Outputs:
    extensions/sd-prompt-composer/group_tags/personal_xlsx.yaml
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

import yaml
import openpyxl

_TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
_SCRIPT_DIR = os.path.dirname(_TOOLS_DIR)
for _p in (_SCRIPT_DIR, _TOOLS_DIR):
    if _p not in sys.path:
        sys.path.insert(0, _p)

DRIVE_DIR = (
    "/Users/hiori222/Library/CloudStorage/GoogleDrive-kanata270@gmail.com/マイドライブ/マイドライブ"
)
SECTION_NAME = "個人辞書 (Google Drive xlsx)"

# Order matters: files appear in this order in the final YAML.
FILES = [
    "1．人物指定（人数、性別、年齢）.xlsx",
    "2．準主体（職業・社会的立場・階級）.xlsx",
    "3．レーティング.xlsx",
    "4．表情、雰囲気.xlsx",
    "5．目の色、特徴.xlsx",
    "6．口、顔、アクセサリー.xlsx",
    "7．髪型、ヘアアクセサリー.xlsx",
    "9．肌、体型.xlsx",
    "10．服装、アクセサリー、小物.xlsx",
    "11．ポーズ、視線、向き.xlsx",
    "12．SEX.xlsx",
    "13．構図、アングル、視点.xlsx",
    "14．背景、環境.xlsx",
    "20．ネガティブ.xlsx",
    "21．ファンタジー.xlsx",
    "prompt_2025_12_28.xlsx",
    "pronpt.xlsx",
    "プロンプトまとめ.xlsx",
    "髪型.xlsx",
]

# Map original filename to a clean category label for the YAML.
FILE_CATEGORY = {
    "1．人物指定（人数、性別、年齢）.xlsx": "1. 人物指定",
    "2．準主体（職業・社会的立場・階級）.xlsx": "2. 準主体（職業/階級）",
    "3．レーティング.xlsx": "3. レーティング",
    "4．表情、雰囲気.xlsx": "4. 表情・雰囲気",
    "5．目の色、特徴.xlsx": "5. 目の色・特徴",
    "6．口、顔、アクセサリー.xlsx": "6. 口・顔・アクセサリー",
    "7．髪型、ヘアアクセサリー.xlsx": "7. 髪型・ヘアアクセサリー",
    "9．肌、体型.xlsx": "9. 肌・体型",
    "10．服装、アクセサリー、小物.xlsx": "10. 服装・小物",
    "11．ポーズ、視線、向き.xlsx": "11. ポーズ・視線・向き",
    "12．SEX.xlsx": "12. SEX",
    "13．構図、アングル、視点.xlsx": "13. 構図・アングル・視点",
    "14．背景、環境.xlsx": "14. 背景・環境",
    "20．ネガティブ.xlsx": "20. ネガティブ",
    "21．ファンタジー.xlsx": "21. ファンタジー",
    "prompt_2025_12_28.xlsx": "プロンプト2025-12-28",
    "pronpt.xlsx": "pronpt",
    "プロンプトまとめ.xlsx": "プロンプトまとめ",
    "髪型.xlsx": "髪型(別冊)",
}

# Sheets to skip entirely (empty placeholders, prose-only TOC sheets, etc.).
SKIP_SHEETS = {
    "Sheet1",
    "Sheet2",
    "Sheet3",
    "Sheet4",
    "Sheet6",  # in プロンプトまとめ.xlsx this is prose
}
ALLOW_SHEET_OVERRIDE = {
    # Some workbooks legitimately use these sheet names; allow them through.
    ("pronpt.xlsx", "Sheet1"),
    ("pronpt.xlsx", "Sheet2"),
    ("pronpt.xlsx", "Sheet3"),
    ("pronpt.xlsx", "Sheet5"),
    ("pronpt.xlsx", "Sheet6"),
    ("pronpt.xlsx", "Sheet7"),
    ("pronpt.xlsx", "Sheet8"),
    ("pronpt.xlsx", "Sheet9"),
    ("髪型.xlsx", "Sheet6"),
    ("プロンプトまとめ.xlsx", "Sheet3"),
    ("プロンプトまとめ.xlsx", "Sheet6"),
}

ENG_TAG_RE = re.compile(r"^[a-zA-Z0-9 _\-,'\.!?():/<>★◆●▼▶]+$")

# Pattern A: ``        english_tag: 日本語``
SINGLE_CELL_PAIR_RE = re.compile(r"^\s*([^:：]+?)\s*[:：]\s*(.+?)\s*$")

# Pattern B (loose): ``english_tag　 日本語`` (separated by Japanese space)
SPACE_SEP_PAIR_RE = re.compile(r"^([a-zA-Z0-9 _\-,'\.!?():/<>]+?)[\s\u3000]+([^a-zA-Z0-9].+)$")

# Pattern C: prose marker for sheet headings
SKIP_LINE_RE = re.compile(r"^(#|\[|＜|<|●|■|□|【|『)")


def _clean_en(text: str) -> str:
    text = text.strip().strip("`")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = re.sub(r"\s+", " ", text)
    text = text.strip(",.;:!?")
    # remove emphasis weights like ((tag)) — keep inner content
    if text.startswith("((") and text.endswith("))"):
        text = text[2:-2].strip()
    if text.startswith("(") and text.endswith(")"):
        # Preserve "(sex)" suffix but drop wrapping parens for "(tag)"
        inner = text[1:-1].strip()
        if " " in inner or ":" in inner:
            text = inner
    return text


def _clean_jp(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    text = text.strip("：:、,.;★◆●▼▶✔ ")
    # cut off long explanations after a separator
    for sep in ("。", "（", "(", "/"):
        if sep in text and len(text.split(sep, 1)[0]) >= 2:
            head = text.split(sep, 1)[0].strip()
            if 1 <= len(head) <= 40:
                text = head
                break
    return text


def _is_useful_english(text: str) -> bool:
    if not text or len(text) > 120:
        return False
    if "○" in text:
        return False
    if "<lora:" in text.lower():
        return False
    if not re.search(r"[a-zA-Z]", text):
        return False
    return bool(ENG_TAG_RE.match(text))


def _looks_jp(text: str) -> bool:
    return bool(re.search(r"[ぁ-んァ-ン一-龥]", text))


def _strip_lead_marker(text: str) -> str:
    """Strip leading non-content markers like ✔, ★, ◆, * before parsing."""
    return re.sub(r"^[\s✔★◆●▼▶◯○※]+", "", text).strip()


def parse_single_cell(text: str) -> List[Tuple[str, str]]:
    """Parse a single cell that may contain ``en_tag: 日本語``."""
    text = _strip_lead_marker(text)
    if not text or SKIP_LINE_RE.match(text):
        return []
    out: List[Tuple[str, str]] = []
    m = SINGLE_CELL_PAIR_RE.match(text)
    if m:
        en = _clean_en(m.group(1))
        jp_raw = m.group(2).strip()
        if not _is_useful_english(en):
            return []
        jp = _clean_jp(jp_raw)
        if not jp or not _looks_jp(jp):
            return []
        out.append((en, jp))
        return out
    m = SPACE_SEP_PAIR_RE.match(text)
    if m:
        en = _clean_en(m.group(1))
        jp = _clean_jp(m.group(2))
        if _is_useful_english(en) and _looks_jp(jp):
            out.append((en, jp))
    return out


def parse_row_cells(cells: List[str]) -> List[Tuple[str, str]]:
    """Parse a row of cells where one column may hold English, another Japanese."""
    if not cells:
        return []
    # Common: 4-column [区分, 項目, 表現内容, プロンプト] → en is last, jp is 3rd
    if len(cells) >= 4 and cells[3] and cells[2]:
        en_candidate = cells[3].strip()
        jp_candidate = cells[2].strip()
        en = _clean_en(en_candidate)
        jp = _clean_jp(jp_candidate)
        if _is_useful_english(en) and _looks_jp(jp):
            return [(en, jp)]
    # 3-column [_, english, japanese]
    if len(cells) >= 3 and cells[1] and cells[2]:
        en = _clean_en(cells[1].strip())
        jp = _clean_jp(cells[2].strip())
        if _is_useful_english(en) and _looks_jp(jp):
            return [(en, jp)]
    # Heuristic: pick any cell that looks English and any that looks Japanese
    en_idx = -1
    jp_idx = -1
    for i, c in enumerate(cells):
        if not c:
            continue
        cs = c.strip()
        if en_idx == -1 and _is_useful_english(cs):
            en_idx = i
        elif jp_idx == -1 and _looks_jp(cs) and not _is_useful_english(cs):
            jp_idx = i
    if en_idx != -1 and jp_idx != -1:
        en = _clean_en(cells[en_idx].strip())
        jp = _clean_jp(cells[jp_idx].strip())
        if _is_useful_english(en) and _looks_jp(jp):
            return [(en, jp)]
    return []


def parse_sheet(ws) -> Dict[str, str]:
    """Return {english_tag: japanese} extracted from one worksheet."""
    out: Dict[str, str] = {}
    # Limit iteration to first 26 cols, all rows.
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Pattern A: single-column row (most other cells empty)
        non_empty = [c for c in cells if c]
        if len(non_empty) == 1:
            for en, jp in parse_single_cell(non_empty[0]):
                out.setdefault(en, jp)
            continue
        # Pattern B: multi-column row
        pairs = parse_row_cells(cells)
        for en, jp in pairs:
            out.setdefault(en, jp)
        # Always also try the first cell as a single-cell pair in case it
        # contains ``en: jp`` even when other cells happen to be filled.
        if cells and cells[0]:
            for en, jp in parse_single_cell(cells[0]):
                out.setdefault(en, jp)
    return out


def build_yaml(per_file: List[Tuple[str, Dict[str, Dict[str, str]]]]) -> List[Dict]:
    section = {"name": SECTION_NAME, "categories": []}
    for cat_name, groups in per_file:
        if not groups:
            continue
        cat_obj = {"name": cat_name, "groups": []}
        for group_name, tags in groups.items():
            if not tags:
                continue
            cat_obj["groups"].append(
                {
                    "name": group_name,
                    "tags": dict(sorted(tags.items(), key=lambda kv: kv[0].lower())),
                }
            )
        if cat_obj["groups"]:
            section["categories"].append(cat_obj)
    return [section]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dir", default=DRIVE_DIR)
    parser.add_argument(
        "--out",
        default=os.path.join(_SCRIPT_DIR, "..", "group_tags", "personal_xlsx.yaml"),
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    per_file: List[Tuple[str, Dict[str, Dict[str, str]]]] = []
    grand_total = 0
    for fname in FILES:
        path = os.path.join(args.dir, fname)
        if not os.path.isfile(path):
            print(f"[skip] {fname}: not found")
            continue
        cat_label = FILE_CATEGORY.get(fname, os.path.splitext(fname)[0])
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            print(f"[error] {fname}: {e}")
            continue
        groups: Dict[str, Dict[str, str]] = {}
        for sn in wb.sheetnames:
            if (fname, sn) not in ALLOW_SHEET_OVERRIDE and sn in SKIP_SHEETS:
                continue
            ws = wb[sn]
            tags = parse_sheet(ws)
            if not tags:
                continue
            groups[sn.strip()] = tags
        wb.close()
        n = sum(len(v) for v in groups.values())
        print(f"[{fname}] -> '{cat_label}': groups={len(groups)} pairs={n}")
        grand_total += n
        per_file.append((cat_label, groups))

    print(f"[personal-xlsx] grand total pairs: {grand_total}")

    if args.dry_run:
        print("[dry-run] not writing YAML")
        return

    data = build_yaml(per_file)
    out_path = os.path.abspath(args.out)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(
            data,
            f,
            allow_unicode=True,
            sort_keys=False,
            default_flow_style=False,
            width=120,
        )
    print(f"[personal-xlsx] wrote: {out_path}")


if __name__ == "__main__":
    main()
